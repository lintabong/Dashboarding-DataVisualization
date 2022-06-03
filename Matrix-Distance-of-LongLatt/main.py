from openpyxl   import load_workbook
from openpyxl   import Workbook
from math       import *
from numpy      import *

def haversine(coord1, coord2):
    R = 6372800  # Earth radius in meters
    lat1, lon1 = coord1
    lat2, lon2 = coord2

    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dlambda = radians(lon2 - lon1)

    a = sin(dphi / 2) ** 2 + \
        cos(phi1) * cos(phi2) * sin(dlambda / 2) ** 2

    return 2 * R * atan2(sqrt(a), sqrt(1 - a))

export_wb   = Workbook()
workbook    = load_workbook("matrix.xlsx")
sheet       = workbook.active
sh_export   = export_wb.active
print(workbook.sheetnames)

harb = {}
nama, latt, long = '', 0, 0
x, y, z          = '', 0, 0
values           = ['V', 'W', 'X']

for i in range(4, 21):
    for o, val in enumerate(values):
        if o == 0:
            x = val
        if o == 1:
            y = val
        if o == 2:
            z = val
    nama = sheet[str(x) + str(i)].value
    latt = sheet[str(y) + str(i)].value
    long = sheet[str(z) + str(i)].value

    harb[nama] = (latt, long)

print(harb)

matrix_distance = zeros([len(harb), len(harb)])

y = 0
for i, o in harb.items():
    print(i)
    x = 0
    for city, coord in harb.items():
        distance = haversine(harb[i], coord)
        print(i, city, distance)
        matrix_distance[x, y] = distance
        x += 1
    y += 1

print(matrix_distance)

a = []
for i in range(67, 84):
    a.append(chr(i))
print(a)

for i, o in enumerate(a):
    for p in range(4, 21):
        sh_export[str(o) + str(p)] = matrix_distance[i, p-4]

export_wb.save("export.xlsx")
